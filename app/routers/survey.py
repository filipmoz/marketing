"""
Survey router for collecting survey responses
"""
from fastapi import APIRouter, Depends, HTTPException, UploadFile, File, Request
from sqlalchemy.orm import Session
from typing import List

from app.database import get_db
from app.survey_models import SurveyResponse
from app.survey_schemas import SurveyResponseCreate, SurveyResponse as SurveyResponseSchema
from app.routers.survey_export import router as export_router
from app.auth import get_current_admin

router = APIRouter()
router.include_router(export_router, prefix="/export", tags=["Export"])

@router.post("/submit", response_model=SurveyResponseSchema)
async def submit_survey(
    response: SurveyResponseCreate,
    db: Session = Depends(get_db)
):
    """
    Submit a survey response
    """
    try:
        db_response = SurveyResponse(
            q1_worried_global_warming=response.q1_worried_global_warming,
            q2_global_warming_threat=response.q2_global_warming_threat,
            q3_british_use_too_much_petrol=response.q3_british_use_too_much_petrol,
            q4_look_petrol_substitutes=response.q4_look_petrol_substitutes,
            q5_petrol_prices_too_high=response.q5_petrol_prices_too_high,
            q6_high_prices_impact_cars=response.q6_high_prices_impact_cars,
            personality_novelist=response.personality_novelist,
            personality_innovator=response.personality_innovator,
            personality_trendsetter=response.personality_trendsetter,
            personality_forerunner=response.personality_forerunner,
            personality_mainstreamer=response.personality_mainstreamer,
            personality_classic=response.personality_classic,
            gender=response.gender,
            marital_status=response.marital_status,
            age_category=response.age_category
        )
        
        db.add(db_response)
        db.commit()
        db.refresh(db_response)
        
        return SurveyResponseSchema(
            id=db_response.id,
            submitted_at=db_response.submitted_at,
            q1_worried_global_warming=db_response.q1_worried_global_warming,
            q2_global_warming_threat=db_response.q2_global_warming_threat,
            q3_british_use_too_much_petrol=db_response.q3_british_use_too_much_petrol,
            q4_look_petrol_substitutes=db_response.q4_look_petrol_substitutes,
            q5_petrol_prices_too_high=db_response.q5_petrol_prices_too_high,
            q6_high_prices_impact_cars=db_response.q6_high_prices_impact_cars,
            personality_novelist=db_response.personality_novelist,
            personality_innovator=db_response.personality_innovator,
            personality_trendsetter=db_response.personality_trendsetter,
            personality_forerunner=db_response.personality_forerunner,
            personality_mainstreamer=db_response.personality_mainstreamer,
            personality_classic=db_response.personality_classic,
            gender=db_response.gender,
            marital_status=db_response.marital_status,
            age_category=db_response.age_category
        )
        
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error submitting survey: {str(e)}")

@router.get("/responses", response_model=List[SurveyResponseSchema])
async def get_all_responses(
    skip: int = 0,
    limit: int = 1000,
    db: Session = Depends(get_db),
    admin: dict = Depends(get_current_admin)
):
    """
    Get all survey responses (for admin/tutor)
    """
    responses = db.query(SurveyResponse).offset(skip).limit(limit).all()
    return [
        SurveyResponseSchema(
            id=r.id,
            submitted_at=r.submitted_at,
            q1_worried_global_warming=r.q1_worried_global_warming,
            q2_global_warming_threat=r.q2_global_warming_threat,
            q3_british_use_too_much_petrol=r.q3_british_use_too_much_petrol,
            q4_look_petrol_substitutes=r.q4_look_petrol_substitutes,
            q5_petrol_prices_too_high=r.q5_petrol_prices_too_high,
            q6_high_prices_impact_cars=r.q6_high_prices_impact_cars,
            personality_novelist=r.personality_novelist,
            personality_innovator=r.personality_innovator,
            personality_trendsetter=r.personality_trendsetter,
            personality_forerunner=r.personality_forerunner,
            personality_mainstreamer=r.personality_mainstreamer,
            personality_classic=r.personality_classic,
            gender=r.gender,
            marital_status=r.marital_status,
            age_category=r.age_category
        )
        for r in responses
    ]

@router.put("/{response_id}/demographics")
async def update_demographics(
    response_id: int,
    gender: str = None,
    marital_status: str = None,
    age_category: str = None,
    db: Session = Depends(get_db)
):
    """
    Update demographics for a survey response
    """
    response = db.query(SurveyResponse).filter(SurveyResponse.id == response_id).first()
    if not response:
        raise HTTPException(status_code=404, detail="Survey response not found")
    
    if gender:
        response.gender = gender
    if marital_status:
        response.marital_status = marital_status
    if age_category:
        response.age_category = age_category
    
    db.commit()
    db.refresh(response)
    
    return SurveyResponseSchema(
        id=response.id,
        submitted_at=response.submitted_at,
        q1_worried_global_warming=response.q1_worried_global_warming,
        q2_global_warming_threat=response.q2_global_warming_threat,
        q3_british_use_too_much_petrol=response.q3_british_use_too_much_petrol,
        q4_look_petrol_substitutes=response.q4_look_petrol_substitutes,
        q5_petrol_prices_too_high=response.q5_petrol_prices_too_high,
        q6_high_prices_impact_cars=response.q6_high_prices_impact_cars,
        personality_novelist=response.personality_novelist,
        personality_innovator=response.personality_innovator,
        personality_trendsetter=response.personality_trendsetter,
        personality_forerunner=response.personality_forerunner,
        personality_mainstreamer=response.personality_mainstreamer,
        personality_classic=response.personality_classic,
        gender=response.gender,
        marital_status=response.marital_status,
        age_category=response.age_category
    )

@router.delete("/clear-all")
async def clear_all_responses(
    db: Session = Depends(get_db),
    admin: dict = Depends(get_current_admin)
):
    """
    Delete all survey responses from the database
    """
    try:
        count = db.query(SurveyResponse).delete()
        db.commit()
        return {"message": f"Successfully deleted {count} survey responses", "deleted_count": count}
    except Exception as e:
        db.rollback()
        raise HTTPException(status_code=500, detail=f"Error clearing database: {str(e)}")

@router.post("/import-excel")
async def import_from_excel(
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
    admin: dict = Depends(get_current_admin)
):
    """
    Import survey responses from Excel file (backup/restore)
    """
    from openpyxl import load_workbook
    from io import BytesIO
    from datetime import datetime
    
    try:
        # Read Excel file
        contents = await file.read()
        workbook = load_workbook(BytesIO(contents))
        
        # Get the "Survey Data" sheet
        if "Survey Data" not in workbook.sheetnames:
            raise HTTPException(status_code=400, detail="Excel file must contain 'Survey Data' sheet")
        
        sheet = workbook["Survey Data"]
        
        # Find header row
        headers = []
        header_row = None
        for row_idx, row in enumerate(sheet.iter_rows(min_row=1, max_row=10, values_only=True), 1):
            if row and any("ID" in str(cell).upper() or "Q1" in str(cell) for cell in row if cell):
                headers = [str(cell).strip() if cell else "" for cell in row]
                header_row = row_idx
                break
        
        if not headers:
            raise HTTPException(status_code=400, detail="Could not find header row in Excel file")
        
        # Map Excel column names to database fields (handle variations)
        column_map = {
            "ID": None,  # Skip ID, will be auto-generated
            "Submitted_At": None,  # Will use current time
            "Submitted At": None,  # Will use current time
            "Q1_Worried_Global_Warming": "q1_worried_global_warming",
            "Q2_Global_Warming_Threat": "q2_global_warming_threat",
            "Q3_British_Use_Too_Much_Petrol": "q3_british_use_too_much_petrol",
            "Q4_Look_Petrol_Substitutes": "q4_look_petrol_substitutes",
            "Q5_Petrol_Prices_Too_High": "q5_petrol_prices_too_high",
            "Q6_High_Prices_Impact_Cars": "q6_high_prices_impact_cars",
            "Personality_Novelist": "personality_novelist",
            "Personality_Innovator": "personality_innovator",
            "Personality_Trendsetter": "personality_trendsetter",
            "Personality_Forerunner": "personality_forerunner",
            "Personality_Mainstreamer": "personality_mainstreamer",
            "Personality_Classic": "personality_classic",
            "Gender": "gender",
            "Marital_Status": "marital_status",
            "Age_Category": "age_category",
        }
        
        # Create column index map (normalize headers)
        col_indices = {}
        for idx, header in enumerate(headers):
            header_clean = str(header).strip()
            if header_clean in column_map:
                db_field = column_map[header_clean]
                if db_field:  # Only map if not None
                    col_indices[db_field] = idx
        
        # Import data rows
        imported = 0
        skipped = 0
        
        for row_idx, row in enumerate(sheet.iter_rows(min_row=header_row + 1, values_only=True), header_row + 1):
            # Skip empty rows
            if not any(cell for cell in row if cell):
                continue
            
            # Extract data
            response_data = {}
            for db_field, col_idx in col_indices.items():
                if col_idx is not None and col_idx < len(row):
                    value = row[col_idx]
                    if value is not None:
                        # Convert to appropriate type
                        if db_field.startswith(('q', 'personality')):
                            try:
                                response_data[db_field] = int(value)
                            except (ValueError, TypeError):
                                skipped += 1
                                break
                        else:
                            response_data[db_field] = str(value).strip()
            
            # Validate required fields
            required_fields = [
                'q1_worried_global_warming', 'q2_global_warming_threat',
                'q3_british_use_too_much_petrol', 'q4_look_petrol_substitutes',
                'q5_petrol_prices_too_high', 'q6_high_prices_impact_cars',
                'personality_novelist', 'personality_innovator', 'personality_trendsetter',
                'personality_forerunner', 'personality_mainstreamer', 'personality_classic'
            ]
            
            if all(field in response_data for field in required_fields):
                # Create response (demographics optional)
                response = SurveyResponse(**response_data)
                db.add(response)
                imported += 1
            else:
                skipped += 1
        
        db.commit()
        
        return {
            "message": f"Successfully imported {imported} survey responses",
            "imported_count": imported,
            "skipped_count": skipped
        }
        
    except HTTPException:
        raise
    except Exception as e:
        db.rollback()
        raise HTTPException(status_code=500, detail=f"Error importing Excel file: {str(e)}")

@router.get("/stats")
async def get_survey_stats(
    db: Session = Depends(get_db),
    admin: dict = Depends(get_current_admin)
):
    """
    Get statistics about survey responses
    """
    total = db.query(SurveyResponse).count()
    
    # Gender distribution
    male_count = db.query(SurveyResponse).filter(SurveyResponse.gender == "Male").count()
    female_count = db.query(SurveyResponse).filter(SurveyResponse.gender == "Female").count()
    
    # Age distribution
    age_18_34 = db.query(SurveyResponse).filter(SurveyResponse.age_category == "18 to 34").count()
    age_35_65 = db.query(SurveyResponse).filter(SurveyResponse.age_category == "35 to 65").count()
    age_65_plus = db.query(SurveyResponse).filter(SurveyResponse.age_category == "65 and older").count()
    
    # Marital status
    married = db.query(SurveyResponse).filter(SurveyResponse.marital_status == "Married").count()
    unmarried = db.query(SurveyResponse).filter(SurveyResponse.marital_status == "Unmarried").count()
    
    return {
        "total_responses": total,
        "gender": {
            "male": male_count,
            "female": female_count
        },
        "age_category": {
            "18_to_34": age_18_34,
            "35_to_65": age_35_65,
            "65_and_older": age_65_plus
        },
        "marital_status": {
            "married": married,
            "unmarried": unmarried
        }
    }

