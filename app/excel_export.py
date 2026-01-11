"""
Excel Export Service for generating .xls files
"""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from typing import List, Dict, Any, Optional
from datetime import datetime
import json
from io import BytesIO

class ExcelExporter:
    """Service for exporting research data to Excel format"""
    
    def __init__(self):
        self.workbook = None
        self.worksheet = None
    
    def create_workbook(self):
        """Create a new workbook"""
        self.workbook = Workbook()
        self.worksheet = self.workbook.active
        self.worksheet.title = "Research Data"
    
    def add_header_row(self, headers: List[str]):
        """Add header row with styling"""
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=12)
        
        for col_num, header in enumerate(headers, 1):
            cell = self.worksheet.cell(row=1, column=col_num)
            cell.value = header
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
        
        # Freeze header row
        self.worksheet.freeze_panes = "A2"
    
    def add_data_rows(self, data: List[Dict[str, Any]]):
        """Add data rows to the worksheet"""
        headers = ["ID", "URL", "Title", "Category", "Content Preview", "Collected At", "Status"]
        
        # Add header if not already added
        if self.worksheet.max_row == 1 and self.worksheet.cell(1, 1).value is None:
            self.add_header_row(headers)
        
        # Add data rows
        for row_num, item in enumerate(data, 2):
            # ID
            self.worksheet.cell(row=row_num, column=1, value=item.get('id'))
            
            # URL
            url_cell = self.worksheet.cell(row=row_num, column=2, value=item.get('url'))
            url_cell.hyperlink = item.get('url')
            url_cell.font = Font(color="0563C1", underline="single")
            
            # Title
            self.worksheet.cell(row=row_num, column=3, value=item.get('title'))
            
            # Category
            self.worksheet.cell(row=row_num, column=4, value=item.get('category'))
            
            # Content Preview (first 200 characters)
            content = item.get('content', '')
            content_preview = content[:200] + "..." if len(content) > 200 else content
            content_cell = self.worksheet.cell(row=row_num, column=5, value=content_preview)
            content_cell.alignment = Alignment(wrap_text=True, vertical="top")
            
            # Collected At
            collected_at = item.get('collected_at')
            if isinstance(collected_at, str):
                collected_at_value = collected_at
            elif hasattr(collected_at, 'isoformat'):
                collected_at_value = collected_at.isoformat()
            else:
                collected_at_value = str(collected_at) if collected_at else ""
            self.worksheet.cell(row=row_num, column=6, value=collected_at_value)
            
            # Status
            status_cell = self.worksheet.cell(row=row_num, column=7, value=item.get('status'))
            # Color code status
            if item.get('status') == 'collected':
                status_cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
            elif item.get('status') == 'error':
                status_cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
        
        # Auto-adjust column widths
        self._auto_adjust_columns()
    
    def _auto_adjust_columns(self):
        """Auto-adjust column widths"""
        column_widths = {
            'A': 8,   # ID
            'B': 40,  # URL
            'C': 30,  # Title
            'D': 15,  # Category
            'E': 50,  # Content Preview
            'F': 20,  # Collected At
            'G': 12   # Status
        }
        
        for col_letter, width in column_widths.items():
            self.worksheet.column_dimensions[col_letter].width = width
    
    def export_to_bytes(self, data: List[Dict[str, Any]]) -> BytesIO:
        """
        Export data to Excel file in memory
        
        Args:
            data: List of research data dictionaries
            
        Returns:
            BytesIO object containing the Excel file
        """
        self.create_workbook()
        self.add_data_rows(data)
        
        # Save to BytesIO
        output = BytesIO()
        self.workbook.save(output)
        output.seek(0)
        return output
    
    def export_to_file(self, data: List[Dict[str, Any]], filename: str):
        """
        Export data to Excel file on disk
        
        Args:
            data: List of research data dictionaries
            filename: Output filename
        """
        self.create_workbook()
        self.add_data_rows(data)
        self.workbook.save(filename)

