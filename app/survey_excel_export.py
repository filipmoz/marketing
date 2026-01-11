"""
Excel Export Service for Survey Data - Code Book Format
Formatted for statistical analysis (crosstabs, pivot tables, etc.)
Includes advanced charts and visualizations
"""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.chart import BarChart, PieChart, LineChart, Reference
from openpyxl.chart.label import DataLabelList
from openpyxl.utils import get_column_letter
from typing import List
from datetime import datetime
from io import BytesIO
from collections import Counter
import numpy as np
try:
    from scipy import stats
    SCIPY_AVAILABLE = True
except ImportError:
    SCIPY_AVAILABLE = False
    # Fallback: basic statistical functions
    import statistics

class SurveyExcelExporter:
    """Service for exporting survey data to Excel in code book format"""
    
    def __init__(self):
        self.workbook = None
        self.data_sheet = None
        self.helper_ranges = {}
    
    def create_workbook(self):
        """Create a new workbook with data sheet"""
        self.workbook = Workbook()
        self.data_sheet = self.workbook.active
        self.data_sheet.title = "Survey Data"
    
    def _reorder_sheets(self, sheet_order: List[str]):
        """Reorder sheets in the workbook to match the desired order"""
        # Create a mapping of sheet titles to sheet objects
        sheet_dict = {sheet.title: sheet for sheet in self.workbook.worksheets}
        
        # Reorder by moving sheets to their desired positions
        # openpyxl uses _sheets list internally for ordering
        ordered_sheets = []
        for sheet_name in sheet_order:
            if sheet_name in sheet_dict:
                ordered_sheets.append(sheet_dict[sheet_name])
        
        # Replace the internal _sheets list
        self.workbook._sheets = ordered_sheets
    
    def _add_survey_footer(self, sheet):
        """Add survey link and credentials footer at the bottom of a sheet"""
        # Find the last row with data
        max_row = sheet.max_row
        footer_row = max_row + 3  # Add some spacing
        
        # Add separator line
        sheet.cell(footer_row, 1, "").border = Border(
            top=Side(style='thin', color='CCCCCC')
        )
        for col in range(2, 7):
            sheet.cell(footer_row, col, "").border = Border(
                top=Side(style='thin', color='CCCCCC')
            )
        
        footer_row += 2
        
        # Add survey link and credentials
        link_cell = sheet.cell(footer_row, 1)
        link_cell.value = "Survey Link: https://filip.kcn.pl"
        link_cell.font = Font(size=10, color="0066CC", underline="single")
        link_cell.hyperlink = "https://filip.kcn.pl"
        
        # Add GitHub repository link next to the survey link
        github_cell = sheet.cell(footer_row, 3)
        github_cell.value = "https://github.com/filipmoz/marketing"
        github_cell.font = Font(size=10, color="0066CC", underline="single")
        github_cell.hyperlink = "https://github.com/filipmoz/marketing"

        cred_cell = sheet.cell(footer_row, 2, "Username: Survey | Password: Filip")
        cred_cell.font = Font(size=10, italic=True, color="666666")
        
        # Adjust column widths to fit the footer text if needed
        # Ensure columns A and B are wide enough for the footer
        def _ensure_min_width(col_letter, min_width):
            dim = sheet.column_dimensions.get(col_letter)
            current = None
            if dim is not None:
                current = getattr(dim, 'width', None)
            if current is None or current < min_width:
                sheet.column_dimensions[col_letter].width = min_width

        _ensure_min_width('A', 35)
        _ensure_min_width('B', 35)
        _ensure_min_width('C', 50)
    
    def export_survey_data(self, responses: List) -> BytesIO:
        """
        Export survey responses to Excel in code book format
        
        Args:
            responses: List of SurveyResponse objects
            
        Returns:
            BytesIO object containing the Excel file
        """
        self.create_workbook()
        
        # Define headers matching the assessment requirements
        headers = [
            'ID',
            # Attitude Questions
            'Q1_Worried_Global_Warming',
            'Q2_Global_Warming_Threat',
            'Q3_British_Use_Too_Much_Petrol',
            'Q4_Look_Petrol_Substitutes',
            'Q5_Petrol_Prices_Too_High',
            'Q6_High_Prices_Impact_Cars',
            # Personality Types
            'Personality_Novelist',
            'Personality_Innovator',
            'Personality_Trendsetter',
            'Personality_Forerunner',
            'Personality_Mainstreamer',
            'Personality_Classic',
            # Demographics
            'Gender',
            'Marital_Status',
            'Age_Category'
        ]
        
        # Add header row with styling
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=11)
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        for col_num, header in enumerate(headers, 1):
            cell = self.data_sheet.cell(row=1, column=col_num)
            cell.value = header
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = border
        
        # Add data rows
        for row_num, response in enumerate(responses, 2):
            data = [
                response.id,
                response.q1_worried_global_warming,
                response.q2_global_warming_threat,
                response.q3_british_use_too_much_petrol,
                response.q4_look_petrol_substitutes,
                response.q5_petrol_prices_too_high,
                response.q6_high_prices_impact_cars,
                response.personality_novelist,
                response.personality_innovator,
                response.personality_trendsetter,
                response.personality_forerunner,
                response.personality_mainstreamer,
                response.personality_classic,
                response.gender,
                response.marital_status,
                response.age_category
            ]
            
            for col_num, value in enumerate(data, 1):
                cell = self.data_sheet.cell(row=row_num, column=col_num)
                cell.value = value
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.border = border
        
        # Auto-adjust column widths
        column_widths = {
            'A': 8,   # ID
            'B': 12,  # Q1
            'C': 12,  # Q2
            'D': 12,  # Q3
            'E': 12,  # Q4
            'F': 12,  # Q5
            'G': 12,  # Q6
            'H': 12,  # Personality_Novelist
            'I': 12,  # Personality_Innovator
            'J': 12,  # Personality_Trendsetter
            'K': 12,  # Personality_Forerunner
            'L': 12,  # Personality_Mainstreamer
            'M': 12,  # Personality_Classic
            'N': 10,  # Gender
            'O': 12,  # Marital_Status
            'P': 15   # Age_Category
        }
        
        for col_letter, width in column_widths.items():
            self.data_sheet.column_dimensions[col_letter].width = width
        
        # Freeze header row
        self.data_sheet.freeze_panes = "A2"
        
        # Add survey footer to Survey Data sheet
        self._add_survey_footer(self.data_sheet)
        
        # Create Code Book sheet (Seminar 4 format)
        self._create_code_book_sheet()
        
        # Create Helper Data sheet FIRST (needed for Statistical Tests sheet)
        # It will be moved to last position later
        self._create_helper_data_sheet(responses)
        
        # Create Analysis Template sheet for statistical tests (position 3)
        self._create_analysis_template_sheet(responses)
        
        # Create Crosstab sheet for Age Group × Innovator (Question 2.a and 2.b)
        self._create_crosstab_sheet(responses)
        
        # Create Statistical Tests sheet (Questions 3, 4, 5) - uses helper_ranges
        self._create_statistical_tests_sheet(responses)
        
        # Create Summary Statistics sheet
        self._create_summary_sheet(responses)
        
        # Create Charts sheet with visualizations
        self._create_charts_sheet(responses)
        
        # Reorder sheets: Survey Data, Code Book, Analysis Templates, Crosstab, Statistical Tests, Summary, Charts, Helper Data
        sheet_order = ['Survey Data', 'Code Book', 'Analysis Templates', 'Crosstab - Age × Innovator', 'Statistical Tests', 'Summary Statistics', 'Charts & Visualizations', 'Helper Data']
        self._reorder_sheets(sheet_order)
        
        # Save to BytesIO
        output = BytesIO()
        self.workbook.save(output)
        output.seek(0)
        return output
    
    def _create_code_book_sheet(self):
        """Create a code book sheet in Seminar 4 format (Description, Statement, Response)"""
        code_sheet = self.workbook.create_sheet("Code Book")
        
        # Title
        code_sheet.merge_cells('A1:F1')
        title_cell = code_sheet.cell(1, 1, "CODE BOOK - Survey Data")
        title_cell.font = Font(bold=True, size=14)
        title_cell.alignment = Alignment(horizontal="center")
        
        # Headers matching Seminar 4 format
        headers = ['DESCRIPTION', 'STATEMENT', 'RESPONSE', 'DESCRIPTION', 'STATEMENT', 'RESPONSE']
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=11)
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        for col_num, header in enumerate(headers, 1):
            cell = code_sheet.cell(3, col_num)
            cell.value = header
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = border
        
        # Code book entries in Seminar 4 format
        # Left column entries
        left_entries = [
            # Attitude Questions
            ('Q1_Worried_Global_Warming', 'I am worried about global warming', ''),
            ('', 'Very strongly disagree', '1'),
            ('', 'Strongly disagree', '2'),
            ('', 'Disagree', '3'),
            ('', 'Neither disagree nor agree', '4'),
            ('', 'Agree', '5'),
            ('', 'Strongly agree', '6'),
            ('', 'Very strongly agree', '7'),
            ('Q2_Global_Warming_Threat', 'Global warming is a real threat', ''),
            ('', 'Very strongly disagree', '1'),
            ('', 'Strongly disagree', '2'),
            ('', 'Disagree', '3'),
            ('', 'Neither disagree nor agree', '4'),
            ('', 'Agree', '5'),
            ('', 'Strongly agree', '6'),
            ('', 'Very strongly agree', '7'),
            ('Q3_British_Use_Too_Much_Petrol', 'British use too much Petrol', ''),
            ('', 'Very strongly disagree', '1'),
            ('', 'Strongly disagree', '2'),
            ('', 'Disagree', '3'),
            ('', 'Neither disagree nor agree', '4'),
            ('', 'Agree', '5'),
            ('', 'Strongly agree', '6'),
            ('', 'Very strongly agree', '7'),
            ('Q4_Look_Petrol_Substitutes', 'We should be looking for Petrol substitutes', ''),
            ('', 'Very strongly disagree', '1'),
            ('', 'Strongly disagree', '2'),
            ('', 'Disagree', '3'),
            ('', 'Neither disagree nor agree', '4'),
            ('', 'Agree', '5'),
            ('', 'Strongly agree', '6'),
            ('', 'Very strongly agree', '7'),
            ('Q5_Petrol_Prices_Too_High', 'Petrol prices are too high now', ''),
            ('', 'Very strongly disagree', '1'),
            ('', 'Strongly disagree', '2'),
            ('', 'Disagree', '3'),
            ('', 'Neither disagree nor agree', '4'),
            ('', 'Agree', '5'),
            ('', 'Strongly agree', '6'),
            ('', 'Very strongly agree', '7'),
            ('Q6_High_Prices_Impact_Cars', 'High gasoline prices will impact what type of cars are purchased', ''),
            ('', 'Very strongly disagree', '1'),
            ('', 'Strongly disagree', '2'),
            ('', 'Disagree', '3'),
            ('', 'Neither disagree nor agree', '4'),
            ('', 'Agree', '5'),
            ('', 'Strongly agree', '6'),
            ('', 'Very strongly agree', '7'),
        ]
        
        # Right column entries
        right_entries = [
            # Personality Types
            ('Personality_Novelist', 'Very early adopter, risk taker, "way out," "show off"', ''),
            ('', 'Does not describe me at all', '1'),
            ('', '2', '2'),
            ('', '3', '3'),
            ('', '4', '4'),
            ('', '5', '5'),
            ('', '6', '6'),
            ('', 'Describes me perfectly', '7'),
            ('Personality_Innovator', 'Early adopter, less risk taker, likes new technology', ''),
            ('', 'Does not describe me at all', '1'),
            ('', '2', '2'),
            ('', '3', '3'),
            ('', '4', '4'),
            ('', '5', '5'),
            ('', '6', '6'),
            ('', 'Describes me perfectly', '7'),
            ('Personality_Trendsetter', 'Opinion leaders, well off financially and educationally', ''),
            ('', 'Does not describe me at all', '1'),
            ('', '2', '2'),
            ('', '3', '3'),
            ('', '4', '4'),
            ('', '5', '5'),
            ('', '6', '6'),
            ('', 'Describes me perfectly', '7'),
            ('Personality_Forerunner', 'Early majority, respected and fairly well off', ''),
            ('', 'Does not describe me at all', '1'),
            ('', '2', '2'),
            ('', '3', '3'),
            ('', '4', '4'),
            ('', '5', '5'),
            ('', '6', '6'),
            ('', 'Describes me perfectly', '7'),
            ('Personality_Mainstreamer', 'Late majority, "average people"', ''),
            ('', 'Does not describe me at all', '1'),
            ('', '2', '2'),
            ('', '3', '3'),
            ('', '4', '4'),
            ('', '5', '5'),
            ('', '6', '6'),
            ('', 'Describes me perfectly', '7'),
            ('Personality_Classic', 'Laggards who cling to "old" ways', ''),
            ('', 'Does not describe me at all', '1'),
            ('', '2', '2'),
            ('', '3', '3'),
            ('', '4', '4'),
            ('', '5', '5'),
            ('', '6', '6'),
            ('', 'Describes me perfectly', '7'),
            # Demographics
            ('Gender', 'What is your gender?', ''),
            ('', 'Male', '1'),
            ('', 'Female', '2'),
            ('Marital_Status', 'What is your marital status?', ''),
            ('', 'Married', '1'),
            ('', 'Unmarried', '2'),
            ('Age_Category', 'What is your age category?', ''),
            ('', '18 to 34', '1'),
            ('', '35 to 65', '2'),
            ('', '65 and older', '3'),
        ]
        
        # Write left column entries
        row = 4
        for desc, stmt, resp in left_entries:
            code_sheet.cell(row, 1, desc)
            code_sheet.cell(row, 2, stmt)
            code_sheet.cell(row, 3, resp)
            for col in [1, 2, 3]:
                code_sheet.cell(row, col).border = border
            row += 1
        
        # Write right column entries
        row = 4
        for desc, stmt, resp in right_entries:
            code_sheet.cell(row, 4, desc)
            code_sheet.cell(row, 5, stmt)
            code_sheet.cell(row, 6, resp)
            for col in [4, 5, 6]:
                code_sheet.cell(row, col).border = border
            row += 1
        
        # Adjust column widths
        code_sheet.column_dimensions['A'].width = 30
        code_sheet.column_dimensions['B'].width = 50
        code_sheet.column_dimensions['C'].width = 12
        code_sheet.column_dimensions['D'].width = 30
        code_sheet.column_dimensions['E'].width = 50
        code_sheet.column_dimensions['F'].width = 12
        
        # Add survey footer
        self._add_survey_footer(code_sheet)
    
    def _create_summary_sheet(self, responses: List):
        """Create summary statistics sheet"""
        summary_sheet = self.workbook.create_sheet("Summary Statistics")
        
        # Title
        summary_sheet.merge_cells('A1:D1')
        title_cell = summary_sheet.cell(1, 1, "Summary Statistics")
        title_cell.font = Font(bold=True, size=14)
        title_cell.alignment = Alignment(horizontal="center")
        
        row = 3
        
        # Demographics Summary
        summary_sheet.cell(row, 1, "Demographics").font = Font(bold=True, size=12)
        row += 1
        
        # Gender distribution
        summary_sheet.cell(row, 1, "Gender Distribution")
        row += 1
        summary_sheet.cell(row, 1, "Male")
        summary_sheet.cell(row, 2, "=COUNTIF('Survey Data'!N:N,\"Male\")")
        summary_sheet.cell(row, 3, "=B" + str(row) + "/COUNTA('Survey Data'!N:N)*100")
        row += 1
        summary_sheet.cell(row, 1, "Female")
        summary_sheet.cell(row, 2, "=COUNTIF('Survey Data'!N:N,\"Female\")")
        summary_sheet.cell(row, 3, "=B" + str(row) + "/COUNTA('Survey Data'!N:N)*100")
        row += 1
        
        row += 1
        
        # Age distribution
        summary_sheet.cell(row, 1, "Age Distribution")
        row += 1
        summary_sheet.cell(row, 1, "18 to 34")
        summary_sheet.cell(row, 2, "=COUNTIF('Survey Data'!P:P,\"18 to 34\")")
        summary_sheet.cell(row, 3, "=B" + str(row) + "/COUNTA('Survey Data'!P:P)*100")
        row += 1
        summary_sheet.cell(row, 1, "35 to 65")
        summary_sheet.cell(row, 2, "=COUNTIF('Survey Data'!P:P,\"35 to 65\")")
        summary_sheet.cell(row, 3, "=B" + str(row) + "/COUNTA('Survey Data'!P:P)*100")
        row += 1
        summary_sheet.cell(row, 1, "65 and older")
        summary_sheet.cell(row, 2, "=COUNTIF('Survey Data'!P:P,\"65 and older\")")
        summary_sheet.cell(row, 3, "=B" + str(row) + "/COUNTA('Survey Data'!P:P)*100")
        row += 1
        
        row += 1
        
        # Marital Status
        summary_sheet.cell(row, 1, "Marital Status Distribution")
        row += 1
        summary_sheet.cell(row, 1, "Married")
        summary_sheet.cell(row, 2, "=COUNTIF('Survey Data'!O:O,\"Married\")")
        summary_sheet.cell(row, 3, "=B" + str(row) + "/COUNTA('Survey Data'!O:O)*100")
        row += 1
        summary_sheet.cell(row, 1, "Unmarried")
        summary_sheet.cell(row, 2, "=COUNTIF('Survey Data'!O:O,\"Unmarried\")")
        summary_sheet.cell(row, 3, "=B" + str(row) + "/COUNTA('Survey Data'!O:O)*100")
        row += 1
        
        row += 2
        
        # Attitude Questions Summary
        summary_sheet.cell(row, 1, "Attitude Questions - Mean Scores").font = Font(bold=True, size=12)
        row += 1
        summary_sheet.cell(row, 1, "Question")
        summary_sheet.cell(row, 2, "Mean")
        summary_sheet.cell(row, 3, "Min")
        summary_sheet.cell(row, 4, "Max")
        summary_sheet.cell(row, 5, "Std Dev")
        row += 1
        
        questions = [
            ("Q1: Worried about global warming", "B"),
            ("Q2: Global warming is a real threat", "C"),
            ("Q3: British use too much Petrol", "D"),
            ("Q4: Look for Petrol substitutes", "E"),
            ("Q5: Petrol prices too high", "F"),
            ("Q6: High prices impact car purchases", "G"),
        ]
        
        for q_name, col_letter in questions:
            summary_sheet.cell(row, 1, q_name)
            summary_sheet.cell(row, 2, f"=AVERAGE('Survey Data'!{col_letter}:{col_letter})")
            summary_sheet.cell(row, 3, f"=MIN('Survey Data'!{col_letter}:{col_letter})")
            summary_sheet.cell(row, 4, f"=MAX('Survey Data'!{col_letter}:{col_letter})")
            summary_sheet.cell(row, 5, f"=STDEV('Survey Data'!{col_letter}:{col_letter})")
            row += 1
        
        row += 2
        
        # Personality Types Summary
        summary_sheet.cell(row, 1, "Personality Types - Mean Scores").font = Font(bold=True, size=12)
        row += 1
        summary_sheet.cell(row, 1, "Personality Type")
        summary_sheet.cell(row, 2, "Mean")
        summary_sheet.cell(row, 3, "Min")
        summary_sheet.cell(row, 4, "Max")
        row += 1
        
        personalities = [
            ("Novelist", "H"),
            ("Innovator", "I"),
            ("Trendsetter", "J"),
            ("Forerunner", "K"),
            ("Mainstreamer", "L"),
            ("Classic", "M"),
        ]
        
        for p_name, col_letter in personalities:
            summary_sheet.cell(row, 1, p_name)
            summary_sheet.cell(row, 2, f"=AVERAGE('Survey Data'!{col_letter}:{col_letter})")
            summary_sheet.cell(row, 3, f"=MIN('Survey Data'!{col_letter}:{col_letter})")
            summary_sheet.cell(row, 4, f"=MAX('Survey Data'!{col_letter}:{col_letter})")
            row += 1
        
        # Adjust column widths
        summary_sheet.column_dimensions['A'].width = 35
        summary_sheet.column_dimensions['B'].width = 12
        summary_sheet.column_dimensions['C'].width = 12
        summary_sheet.column_dimensions['D'].width = 12
        
        # Add survey footer
        self._add_survey_footer(summary_sheet)
        summary_sheet.column_dimensions['E'].width = 12
    
    def _create_charts_sheet(self, responses: List):
        """Create charts sheet with visualizations using formulas from survey data"""
        charts_sheet = self.workbook.create_sheet("Charts & Visualizations")
        
        # Title
        charts_sheet.merge_cells('A1:J1')
        title_cell = charts_sheet.cell(1, 1, "Data Visualizations")
        title_cell.font = Font(bold=True, size=14, color="FFFFFF")
        title_cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        title_cell.alignment = Alignment(horizontal="center", vertical="center")
        
        row = 3
        max_rows = len(responses) + 1
        
        # 1. Gender Distribution Pie Chart
        charts_sheet.cell(row, 1, "Gender Distribution").font = Font(bold=True, size=12)
        row += 1
        charts_sheet.cell(row, 1, "Gender").font = Font(bold=True)
        charts_sheet.cell(row, 2, "Count").font = Font(bold=True)
        row += 1
        charts_sheet.cell(row, 1, "Female")
        charts_sheet.cell(row, 2, f"=COUNTIF('Survey Data'!N:N,\"Female\")")
        row += 1
        charts_sheet.cell(row, 1, "Male")
        charts_sheet.cell(row, 2, f"=COUNTIF('Survey Data'!N:N,\"Male\")")
        gender_data_end = row
        
        # Create Pie Chart for Gender (smaller to avoid overlap)
        pie_chart = PieChart()
        pie_chart.title = "Gender Distribution"
        pie_chart.width = 15  # Make chart smaller to avoid overlap
        pie_chart.height = 12
        pie_chart.legend = None  # Remove legend
        data = Reference(charts_sheet, min_col=2, min_row=row-2, max_row=gender_data_end)
        cats = Reference(charts_sheet, min_col=1, min_row=row-2, max_row=gender_data_end)
        pie_chart.add_data(data, titles_from_data=False)
        pie_chart.set_categories(cats)
        # Set series title to empty to avoid showing "Column B"
        if pie_chart.series:
            pie_chart.series[0].title = None
        pie_chart.dataLabels = DataLabelList()
        pie_chart.dataLabels.showPercent = True
        pie_chart.dataLabels.showCatName = True  # Show category name
        pie_chart.dataLabels.showVal = True  # Show value
        pie_chart.dataLabels.showSerName = False  # Don't show series name
        charts_sheet.add_chart(pie_chart, "F3")  # Position chart further right to avoid overlap
        
        row = gender_data_end + 20  # More space between charts to avoid overlap
        
        # 2. Age Category Bar Chart
        charts_sheet.cell(row, 1, "Age Category Distribution").font = Font(bold=True, size=12)
        row += 1
        charts_sheet.cell(row, 1, "Age Category").font = Font(bold=True)
        charts_sheet.cell(row, 2, "Count").font = Font(bold=True)
        row += 1
        charts_sheet.cell(row, 1, "18 to 34")
        charts_sheet.cell(row, 2, f"=COUNTIF('Survey Data'!P:P,\"18 to 34\")")
        row += 1
        charts_sheet.cell(row, 1, "35 to 65")
        charts_sheet.cell(row, 2, f"=COUNTIF('Survey Data'!P:P,\"35 to 65\")")
        row += 1
        charts_sheet.cell(row, 1, "65 and older")
        charts_sheet.cell(row, 2, f"=COUNTIF('Survey Data'!P:P,\"65 and older\")")
        age_data_end = row
        
        # Create Bar Chart for Age (bigger)
        bar_chart = BarChart()
        bar_chart.type = "col"
        bar_chart.style = 10
        bar_chart.title = "Age Category Distribution"
        bar_chart.y_axis.title = "Count"
        bar_chart.x_axis.title = "Age Category"
        bar_chart.width = 15
        bar_chart.height = 10
        bar_chart.legend = None  # Remove legend
        data = Reference(charts_sheet, min_col=2, min_row=age_data_end-3, max_row=age_data_end)
        cats = Reference(charts_sheet, min_col=1, min_row=age_data_end-3, max_row=age_data_end)
        bar_chart.add_data(data, titles_from_data=False)
        bar_chart.set_categories(cats)
        # Set series title to empty to avoid showing "Column B"
        if bar_chart.series:
            bar_chart.series[0].title = None
        bar_chart.dataLabels = DataLabelList()
        bar_chart.dataLabels.showVal = True
        bar_chart.dataLabels.showSerName = False  # Don't show series name
        charts_sheet.add_chart(bar_chart, "F" + str(age_data_end - 8))  # Position further right
        
        row = age_data_end + 20  # More space to avoid overlap
        
        # 3. Attitude Questions Mean Scores
        questions = [
            ("Q1", "B", "Worried about Global Warming"),
            ("Q2", "C", "Global Warming is a Threat"),
            ("Q3", "D", "British Use Too Much Petrol"),
            ("Q4", "E", "Look for Petrol Substitutes"),
            ("Q5", "F", "Petrol Prices Too High"),
            ("Q6", "G", "High Prices Impact Cars"),
        ]
        
        chart_start_row = row
        charts_sheet.cell(chart_start_row, 1, "Attitude Questions - Mean Scores").font = Font(bold=True, size=12)
        charts_sheet.cell(chart_start_row + 1, 1, "Question").font = Font(bold=True)
        charts_sheet.cell(chart_start_row + 1, 2, "Mean Score").font = Font(bold=True)
        row = chart_start_row + 2
        
        for q_name, col_letter, q_desc in questions:
            charts_sheet.cell(row, 1, q_name)
            avg_cell = charts_sheet.cell(row, 2, f"=AVERAGE('Survey Data'!{col_letter}:{col_letter})")
            avg_cell.number_format = '0.00'  # Format to 2 decimal places
            row += 1
        
        # Create Line Chart for Attitude Questions (bigger)
        line_chart = LineChart()
        line_chart.title = "Attitude Questions - Mean Scores"
        line_chart.style = 13
        line_chart.y_axis.title = "Mean Score (1-7)"
        line_chart.x_axis.title = "Question"
        line_chart.width = 20  # Make bigger
        line_chart.height = 12
        line_chart.legend = None  # Remove legend
        data = Reference(charts_sheet, min_col=2, min_row=chart_start_row+2, max_row=chart_start_row+1+len(questions))
        cats = Reference(charts_sheet, min_col=1, min_row=chart_start_row+2, max_row=chart_start_row+1+len(questions))
        line_chart.add_data(data, titles_from_data=False)
        line_chart.set_categories(cats)
        # Set series title to empty to avoid showing "Column B"
        if line_chart.series:
            line_chart.series[0].title = None
        line_chart.dataLabels = DataLabelList()
        line_chart.dataLabels.showVal = True
        line_chart.dataLabels.showSerName = False  # Don't show series name
        charts_sheet.add_chart(line_chart, "F" + str(chart_start_row))  # Position further right
        
        row = chart_start_row + len(questions) + 20  # More space to avoid overlap
        
        # 4. Personality Types Comparison
        personalities = [
            ("Novelist", "H"),
            ("Innovator", "I"),
            ("Trendsetter", "J"),
            ("Forerunner", "K"),
            ("Mainstreamer", "L"),
            ("Classic", "M"),
        ]
        
        personality_start_row = row
        charts_sheet.cell(personality_start_row, 1, "Personality Types - Mean Scores").font = Font(bold=True, size=12)
        charts_sheet.cell(personality_start_row + 1, 1, "Personality Type").font = Font(bold=True)
        charts_sheet.cell(personality_start_row + 1, 2, "Mean Score").font = Font(bold=True)
        row = personality_start_row + 2
        
        for p_name, col_letter in personalities:
            charts_sheet.cell(row, 1, p_name)
            avg_cell = charts_sheet.cell(row, 2, f"=AVERAGE('Survey Data'!{col_letter}:{col_letter})")
            avg_cell.number_format = '0.00'  # Format to 2 decimal places
            row += 1
        
        # Create Bar Chart for Personality Types (bigger)
        personality_chart = BarChart()
        personality_chart.type = "col"
        personality_chart.style = 10
        personality_chart.title = "Personality Types - Mean Scores"
        personality_chart.y_axis.title = "Mean Score (1-7)"
        personality_chart.x_axis.title = "Personality Type"
        personality_chart.width = 20  # Make bigger
        personality_chart.height = 12
        personality_chart.legend = None  # Remove legend
        data = Reference(charts_sheet, min_col=2, min_row=personality_start_row+2, max_row=personality_start_row+1+len(personalities))
        cats = Reference(charts_sheet, min_col=1, min_row=personality_start_row+2, max_row=personality_start_row+1+len(personalities))
        personality_chart.add_data(data, titles_from_data=False)
        personality_chart.set_categories(cats)
        # Set series title to empty to avoid showing "Column B"
        if personality_chart.series:
            personality_chart.series[0].title = None
        personality_chart.dataLabels = DataLabelList()
        personality_chart.dataLabels.showVal = True
        personality_chart.dataLabels.showSerName = False  # Don't show series name
        charts_sheet.add_chart(personality_chart, "F" + str(personality_start_row))  # Position further right
        
        row = personality_start_row + len(personalities) + 20  # More space to avoid overlap
        
        # 5. Marital Status Distribution
        marital_start_row = row
        charts_sheet.cell(marital_start_row, 1, "Marital Status Distribution").font = Font(bold=True, size=12)
        charts_sheet.cell(marital_start_row + 1, 1, "Status").font = Font(bold=True)
        charts_sheet.cell(marital_start_row + 1, 2, "Count").font = Font(bold=True)
        row = marital_start_row + 2
        charts_sheet.cell(row, 1, "Married")
        charts_sheet.cell(row, 2, f"=COUNTIF('Survey Data'!O:O,\"Married\")")
        row += 1
        charts_sheet.cell(row, 1, "Unmarried")
        charts_sheet.cell(row, 2, f"=COUNTIF('Survey Data'!O:O,\"Unmarried\")")
        marital_data_end = row
        
        # Create Pie Chart for Marital Status (bigger)
        marital_pie = PieChart()
        marital_pie.title = "Marital Status Distribution"
        marital_pie.width = 20
        marital_pie.height = 15
        marital_pie.legend = None  # Remove legend
        data = Reference(charts_sheet, min_col=2, min_row=marital_start_row+1, max_row=marital_data_end)
        cats = Reference(charts_sheet, min_col=1, min_row=marital_start_row+2, max_row=marital_data_end)
        marital_pie.add_data(data, titles_from_data=False)
        marital_pie.set_categories(cats)
        # Set series title to empty to avoid showing "Column B"
        if marital_pie.series:
            marital_pie.series[0].title = None
        marital_pie.dataLabels = DataLabelList()
        marital_pie.dataLabels.showPercent = True
        marital_pie.dataLabels.showCatName = True  # Show category name
        marital_pie.dataLabels.showVal = True  # Show value
        marital_pie.dataLabels.showSerName = False  # Don't show series name
        charts_sheet.add_chart(marital_pie, "F" + str(marital_start_row))  # Position further right to avoid overlap
        
        # Adjust column widths
        charts_sheet.column_dimensions['A'].width = 30
        charts_sheet.column_dimensions['B'].width = 15
        
        # Add survey footer
        self._add_survey_footer(charts_sheet)
    
    def _create_analysis_template_sheet(self, responses: List):
        """Create analysis template sheet for statistical tests (crosstabs, t-tests, ANOVA, chi-square)"""
        analysis_sheet = self.workbook.create_sheet("Analysis Templates")
        
        # Title
        analysis_sheet.merge_cells('A1:F1')
        title_cell = analysis_sheet.cell(1, 1, "Statistical Analysis Templates & Guidance")
        title_cell.font = Font(bold=True, size=14)
        title_cell.alignment = Alignment(horizontal="center")
        
        row = 3
        
        # Research Aims and Objectives Section (Question 1.a)
        analysis_sheet.cell(row, 1, "1. RESEARCH AIMS AND OBJECTIVES (Question 1.a)").font = Font(bold=True, size=12, color="366092")
        row += 1
        analysis_sheet.cell(row, 1, "Research Topic:")
        analysis_sheet.cell(row, 2, "A prominent car manufacturer is seeking to understand consumer attitudes towards fuel prices, global warming, and alternative fuels.")
        analysis_sheet.merge_cells(f'B{row}:F{row}')
        row += 1
        analysis_sheet.cell(row, 1, "Research Question:")
        analysis_sheet.cell(row, 2, "How do consumer perceptions of global warming, petrol usage, and fuel prices influence their preferences for alternative fuel vehicles?")
        analysis_sheet.merge_cells(f'B{row}:F{row}')
        row += 1
        analysis_sheet.cell(row, 1, "Research Aim:")
        analysis_sheet.cell(row, 2, "To investigate the relationship between consumer perceptions of environmental issues (global warming), fuel consumption patterns (petrol usage), and economic factors (fuel prices) and their preferences for alternative fuel vehicles, in order to inform the car manufacturer's future vehicle development and marketing strategies.")
        analysis_sheet.merge_cells(f'B{row}:F{row}')
        row += 1
        analysis_sheet.cell(row, 1, "Objective 1:")
        analysis_sheet.cell(row, 2, "To examine the extent to which consumer concerns about global warming and environmental issues influence their attitudes towards alternative fuel vehicles.")
        analysis_sheet.merge_cells(f'B{row}:F{row}')
        row += 1
        analysis_sheet.cell(row, 1, "Objective 2:")
        analysis_sheet.cell(row, 2, "To analyze how consumer perceptions of petrol prices and fuel consumption patterns affect their preferences for alternative fuel vehicle options.")
        analysis_sheet.merge_cells(f'B{row}:F{row}')
        row += 2
        
        # Crosstab Template Section
        analysis_sheet.cell(row, 1, "2. CROSSTABULATION TEMPLATE").font = Font(bold=True, size=12, color="366092")
        row += 1
        analysis_sheet.cell(row, 1, "Note:")
        analysis_sheet.cell(row, 2, "A complete crosstabulation for Age Group × Innovator Personality (High/Low) has been created in the 'Crosstab - Age × Innovator' sheet. See that sheet for the actual data and hypothesis testing.")
        analysis_sheet.merge_cells(f'B{row}:F{row}')
        row += 1
        analysis_sheet.cell(row, 1, "Example Formula Template (for reference): Age Group × Innovator Personality (High/Low)")
        row += 1
        analysis_sheet.cell(row, 2, "Low Innovator")
        analysis_sheet.cell(row, 3, "High Innovator")
        analysis_sheet.cell(row, 4, "Total")
        for cell in [analysis_sheet.cell(row, 2), analysis_sheet.cell(row, 3), analysis_sheet.cell(row, 4)]:
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")
        row += 1
        analysis_sheet.cell(row, 1, "18 to 34")
        analysis_sheet.cell(row, 2, "=COUNTIFS('Survey Data'!P:P,\"18 to 34\",'Survey Data'!I:I,\"<5\")")
        analysis_sheet.cell(row, 3, "=COUNTIFS('Survey Data'!P:P,\"18 to 34\",'Survey Data'!I:I,\">=5\")")
        analysis_sheet.cell(row, 4, f"=SUM(B{row}:C{row})")
        row += 1
        analysis_sheet.cell(row, 1, "35 to 65")
        analysis_sheet.cell(row, 2, "=COUNTIFS('Survey Data'!P:P,\"35 to 65\",'Survey Data'!I:I,\"<5\")")
        analysis_sheet.cell(row, 3, "=COUNTIFS('Survey Data'!P:P,\"35 to 65\",'Survey Data'!I:I,\">=5\")")
        analysis_sheet.cell(row, 4, f"=SUM(B{row}:C{row})")
        row += 1
        analysis_sheet.cell(row, 1, "65 and older")
        analysis_sheet.cell(row, 2, "=COUNTIFS('Survey Data'!P:P,\"65 and older\",'Survey Data'!I:I,\"<5\")")
        analysis_sheet.cell(row, 3, "=COUNTIFS('Survey Data'!P:P,\"65 and older\",'Survey Data'!I:I,\">=5\")")
        analysis_sheet.cell(row, 4, f"=SUM(B{row}:C{row})")
        row += 1
        analysis_sheet.cell(row, 1, "Total")
        analysis_sheet.cell(row, 2, f"=SUM(B{row-3}:B{row-1})")
        analysis_sheet.cell(row, 3, f"=SUM(C{row-3}:C{row-1})")
        analysis_sheet.cell(row, 4, f"=SUM(D{row-3}:D{row-1})")
        for cell in [analysis_sheet.cell(row, 1), analysis_sheet.cell(row, 2), analysis_sheet.cell(row, 3), analysis_sheet.cell(row, 4)]:
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")
        row += 2
        
        # Statistical Test Selection Guide
        analysis_sheet.cell(row, 1, "3. STATISTICAL TEST SELECTION GUIDE").font = Font(bold=True, size=12, color="366092")
        row += 1
        analysis_sheet.cell(row, 1, "Test Type")
        analysis_sheet.cell(row, 2, "When to Use")
        analysis_sheet.cell(row, 3, "Variables")
        for cell in [analysis_sheet.cell(row, 1), analysis_sheet.cell(row, 2), analysis_sheet.cell(row, 3)]:
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")
        row += 1
        analysis_sheet.cell(row, 1, "Chi-Square")
        analysis_sheet.cell(row, 2, "Testing association between two categorical variables")
        analysis_sheet.cell(row, 3, "Categorical × Categorical")
        row += 1
        analysis_sheet.cell(row, 1, "T-Test (Independent)")
        analysis_sheet.cell(row, 2, "Comparing means of two groups")
        analysis_sheet.cell(row, 3, "Continuous × Categorical (2 groups)")
        row += 1
        analysis_sheet.cell(row, 1, "T-Test (Paired)")
        analysis_sheet.cell(row, 2, "Comparing means of same group on two variables")
        analysis_sheet.cell(row, 3, "Two continuous variables (same subjects)")
        row += 1
        analysis_sheet.cell(row, 1, "ANOVA")
        analysis_sheet.cell(row, 2, "Comparing means across three or more groups")
        analysis_sheet.cell(row, 3, "Continuous × Categorical (3+ groups)")
        row += 2
        
        # Interpretation Guide
        analysis_sheet.cell(row, 1, "4. INTERPRETATION GUIDE").font = Font(bold=True, size=12, color="366092")
        row += 1
        analysis_sheet.cell(row, 1, "Significance Level (α):")
        analysis_sheet.cell(row, 2, "Typically 0.05 (5%)")
        row += 1
        analysis_sheet.cell(row, 1, "If p-value < 0.05:")
        analysis_sheet.cell(row, 2, "Reject H0 - There is a statistically significant result")
        analysis_sheet.merge_cells(f'B{row}:F{row}')
        row += 1
        analysis_sheet.cell(row, 1, "If p-value ≥ 0.05:")
        analysis_sheet.cell(row, 2, "Fail to reject H0 - No statistically significant result")
        analysis_sheet.merge_cells(f'B{row}:F{row}')
        row += 1
        analysis_sheet.cell(row, 1, "Effect Size:")
        analysis_sheet.cell(row, 2, "Consider practical significance, not just statistical significance")
        analysis_sheet.merge_cells(f'B{row}:F{row}')
        row += 2
        
        # Excel Functions Reference
        analysis_sheet.cell(row, 1, "5. USEFUL EXCEL FUNCTIONS").font = Font(bold=True, size=12, color="366092")
        row += 1
        analysis_sheet.cell(row, 1, "Function")
        analysis_sheet.cell(row, 2, "Purpose")
        analysis_sheet.cell(row, 3, "Example")
        for cell in [analysis_sheet.cell(row, 1), analysis_sheet.cell(row, 2), analysis_sheet.cell(row, 3)]:
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")
        row += 1
        analysis_sheet.cell(row, 1, "COUNTIFS")
        analysis_sheet.cell(row, 2, "Count with multiple criteria")
        # Store as text format to display formula as example
        example_cell = analysis_sheet.cell(row, 3)
        example_cell.value = "COUNTIFS(A:A,\"Male\",B:B,\">5\")"
        example_cell.number_format = '@'  # Text format
        row += 1
        analysis_sheet.cell(row, 1, "AVERAGEIF")
        analysis_sheet.cell(row, 2, "Average with condition")
        example_cell = analysis_sheet.cell(row, 3)
        example_cell.value = "AVERAGEIF(A:A,\"Married\",B:B)"
        example_cell.number_format = '@'  # Text format
        row += 1
        analysis_sheet.cell(row, 1, "AVERAGEIFS")
        analysis_sheet.cell(row, 2, "Average with multiple conditions")
        example_cell = analysis_sheet.cell(row, 3)
        example_cell.value = "AVERAGEIFS(B:B,A:A,\"Female\",C:C,\"18 to 34\")"
        example_cell.number_format = '@'  # Text format
        row += 1
        analysis_sheet.cell(row, 1, "STDEV")
        analysis_sheet.cell(row, 2, "Standard deviation")
        example_cell = analysis_sheet.cell(row, 3)
        example_cell.value = "STDEV(A:A)"
        example_cell.number_format = '@'  # Text format
        row += 1
        analysis_sheet.cell(row, 1, "CHITEST")
        analysis_sheet.cell(row, 2, "Chi-square test p-value")
        example_cell = analysis_sheet.cell(row, 3)
        example_cell.value = "CHITEST(actual_range, expected_range)"
        example_cell.number_format = '@'  # Text format
        row += 1
        analysis_sheet.cell(row, 1, "T.TEST")
        analysis_sheet.cell(row, 2, "T-test p-value")
        example_cell = analysis_sheet.cell(row, 3)
        example_cell.value = "T.TEST(array1, array2, tails, type)"
        example_cell.number_format = '@'  # Text format
        
        # Adjust column widths
        analysis_sheet.column_dimensions['A'].width = 30
        analysis_sheet.column_dimensions['B'].width = 50
        analysis_sheet.column_dimensions['C'].width = 30
        analysis_sheet.column_dimensions['D'].width = 15
        analysis_sheet.column_dimensions['E'].width = 15
        analysis_sheet.column_dimensions['F'].width = 15
        
        # Add survey footer
        self._add_survey_footer(analysis_sheet)

    def _create_crosstab_sheet(self, responses: List):
        """Create crosstab sheet for Age Group × Innovator (Question 2.a and 2.b)"""
        crosstab_sheet = self.workbook.create_sheet("Crosstab - Age × Innovator")
        
        # Title
        crosstab_sheet.merge_cells('A1:E1')
        title_cell = crosstab_sheet.cell(1, 1, "Crosstabulation: Age Group × Innovator Personality (Question 2.a)")
        title_cell.font = Font(bold=True, size=14)
        title_cell.alignment = Alignment(horizontal="center")
        
        row = 3
        
        # Classification explanation
        crosstab_sheet.cell(row, 1, "Innovator Classification:").font = Font(bold=True, size=11)
        row += 1
        crosstab_sheet.cell(row, 1, "Low Innovator:")
        crosstab_sheet.cell(row, 2, "Very strongly disagree (1), Strongly disagree (2), Disagree (3), Neither disagree nor agree (4)")
        crosstab_sheet.merge_cells(f'B{row}:E{row}')
        row += 1
        crosstab_sheet.cell(row, 1, "High Innovator:")
        crosstab_sheet.cell(row, 2, "Agree (5), Strongly agree (6), Very strongly agree (7)")
        crosstab_sheet.merge_cells(f'B{row}:E{row}')
        row += 2
        
        # Create the crosstab table
        # Headers
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=11)
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        crosstab_sheet.cell(row, 1, "Age Group")
        crosstab_sheet.cell(row, 2, "Low Innovator")
        crosstab_sheet.cell(row, 3, "High Innovator")
        crosstab_sheet.cell(row, 4, "Total")
        
        for col in [1, 2, 3, 4]:
            cell = crosstab_sheet.cell(row, col)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = border
        
        row += 1
        
        # Age groups
        age_groups = ["18 to 34", "35 to 65", "65 and older"]
        
        # Use Excel formulas for counts
        start_data_row = row
        for age_group in age_groups:
            crosstab_sheet.cell(row, 1, age_group)
            crosstab_sheet.cell(row, 2, f"=COUNTIFS('Survey Data'!P:P,\"{age_group}\",'Survey Data'!I:I,\"<5\")")
            crosstab_sheet.cell(row, 3, f"=COUNTIFS('Survey Data'!P:P,\"{age_group}\",'Survey Data'!I:I,\">=5\")")
            crosstab_sheet.cell(row, 4, f"=SUM(B{row}:C{row})")
            
            for col in [1, 2, 3, 4]:
                cell = crosstab_sheet.cell(row, col)
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.border = border
            
            row += 1
        
        # Total row with formulas
        crosstab_sheet.cell(row, 1, "Total").font = Font(bold=True)
        crosstab_sheet.cell(row, 2, f"=SUM(B{start_data_row}:B{row-1})").font = Font(bold=True)
        crosstab_sheet.cell(row, 3, f"=SUM(C{start_data_row}:C{row-1})").font = Font(bold=True)
        crosstab_sheet.cell(row, 4, f"=SUM(D{start_data_row}:D{row-1})").font = Font(bold=True)
        
        for col in [1, 2, 3, 4]:
            cell = crosstab_sheet.cell(row, col)
            cell.fill = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = border
        
        row += 3
        
        # Hypothesis Testing Section (Question 2.b)
        crosstab_sheet.cell(row, 1, "Hypothesis Testing (Question 2.b)").font = Font(bold=True, size=12, color="366092")
        row += 1
        
        crosstab_sheet.cell(row, 1, "Null Hypothesis (H0):")
        crosstab_sheet.cell(row, 2, "There is no significant association between Age Group and Innovator personality classification (Low/High).")
        crosstab_sheet.merge_cells(f'B{row}:E{row}')
        row += 1
        
        crosstab_sheet.cell(row, 1, "Alternative Hypothesis (H1):")
        crosstab_sheet.cell(row, 2, "There is a significant association between Age Group and Innovator personality classification (Low/High).")
        crosstab_sheet.merge_cells(f'B{row}:E{row}')
        row += 1
        
        crosstab_sheet.cell(row, 1, "Statistical Test:")
        crosstab_sheet.cell(row, 2, "Chi-Square Test of Independence")
        crosstab_sheet.merge_cells(f'B{row}:E{row}')
        row += 1
        
        crosstab_sheet.cell(row, 1, "Test Rationale:")
        crosstab_sheet.cell(row, 2, "The Chi-Square test is appropriate because both variables (Age Group and Innovator classification) are categorical variables. This test determines if there is a statistically significant association between the two categorical variables.")
        crosstab_sheet.merge_cells(f'B{row}:E{row}')
        row += 1
        
        crosstab_sheet.cell(row, 1, "How to Conduct the Test in Excel:")
        crosstab_sheet.merge_cells(f'B{row}:E{row}')
        row += 1
        crosstab_sheet.cell(row, 1, "Method 1:")
        crosstab_sheet.cell(row, 2, "Use Data Analysis ToolPak: Data > Data Analysis > Chi-Square Test")
        crosstab_sheet.merge_cells(f'B{row}:E{row}')
        row += 1
        crosstab_sheet.cell(row, 1, "Method 2:")
        crosstab_sheet.cell(row, 2, "Use CHITEST() function with observed and expected frequencies")
        crosstab_sheet.merge_cells(f'B{row}:E{row}')
        row += 1
        
        crosstab_sheet.cell(row, 1, "Expected Frequencies (for reference):")
        crosstab_sheet.merge_cells(f'B{row}:E{row}')
        expected_header_row = row
        row += 1
        
        # Expected frequencies using Excel formulas
        crosstab_sheet.cell(row, 1, "Age Group")
        crosstab_sheet.cell(row, 2, "Low Innovator (Expected)")
        crosstab_sheet.cell(row, 3, "High Innovator (Expected)")
        for col in [1, 2, 3]:
            cell = crosstab_sheet.cell(row, col)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
            cell.border = border
        row += 1
        
        # Store where expected frequencies data starts
        expected_data_start_row = row
        
        # Use Excel formulas for expected frequencies: (row total * column total) / grand total
        for i, age_group in enumerate(age_groups):
            data_row = start_data_row + i
            crosstab_sheet.cell(row, 1, age_group)
            # Expected = (Row Total * Column Total) / Grand Total
            crosstab_sheet.cell(row, 2, f"=D{data_row}*B{start_data_row + len(age_groups)}/D{start_data_row + len(age_groups)}")
            crosstab_sheet.cell(row, 3, f"=D{data_row}*C{start_data_row + len(age_groups)}/D{start_data_row + len(age_groups)}")
            for col in [1, 2, 3]:
                cell = crosstab_sheet.cell(row, col)
                cell.border = border
            row += 1
        
        row += 1
        crosstab_sheet.cell(row, 1, "Interpretation Guide:")
        crosstab_sheet.merge_cells(f'B{row}:E{row}')
        row += 1
        crosstab_sheet.cell(row, 1, "If p-value < 0.05:")
        crosstab_sheet.cell(row, 2, "Reject H0 - There is a statistically significant association between Age Group and Innovator personality classification.")
        crosstab_sheet.merge_cells(f'B{row}:E{row}')
        row += 1
        crosstab_sheet.cell(row, 1, "If p-value ≥ 0.05:")
        crosstab_sheet.cell(row, 2, "Fail to reject H0 - There is no statistically significant association between Age Group and Innovator personality classification.")
        crosstab_sheet.merge_cells(f'B{row}:E{row}')
        row += 2
        
        # Chi-Square Test Results using Excel formulas
        crosstab_sheet.cell(row, 1, "Chi-Square Test Results:").font = Font(bold=True, size=11)
        row += 1
        
        # Actually, let's calculate it properly: expected frequencies table starts after "Expected Frequencies" header
        # We'll reference it dynamically by finding the expected frequencies section
        # For now, let's use a simpler approach - reference the expected frequencies we calculated
        # Expected frequencies are in columns B and C, starting after the header row
        
        crosstab_sheet.cell(row, 1, "Degrees of Freedom:")
        crosstab_sheet.cell(row, 2, "=(ROWS(B" + str(start_data_row) + ":B" + str(start_data_row + len(age_groups) - 1) + ")-1)*(COLUMNS(B" + str(start_data_row) + ":C" + str(start_data_row) + ")-1)")
        row += 1
        crosstab_sheet.cell(row, 1, "Significance Level (α):")
        crosstab_sheet.cell(row, 2, "0.05")
        row += 1
        
        # Calculate p-value using CHITEST
        observed_range = f"B{start_data_row}:C{start_data_row + len(age_groups) - 1}"
        expected_range = f"B{expected_data_start_row}:C{expected_data_start_row + len(age_groups) - 1}"
        
        crosstab_sheet.cell(row, 1, "p-value (CHITEST):")
        crosstab_sheet.cell(row, 2, f"=CHITEST({observed_range},{expected_range})")
        row += 1
        crosstab_sheet.cell(row, 1, "Conclusion:").font = Font(bold=True)
        p_value_row = row - 1
        crosstab_sheet.cell(row, 2, f'=IF(B{p_value_row}<0.05,"Reject H0 - There is a statistically significant association","Fail to reject H0 - No statistically significant association")')
        crosstab_sheet.merge_cells(f'B{row}:E{row}')
        
        # Adjust column widths
        crosstab_sheet.column_dimensions['A'].width = 20
        crosstab_sheet.column_dimensions['B'].width = 20
        crosstab_sheet.column_dimensions['C'].width = 20
        crosstab_sheet.column_dimensions['D'].width = 15
        crosstab_sheet.column_dimensions['E'].width = 50
        
        # Add survey footer
        self._add_survey_footer(crosstab_sheet)
    
    def _create_helper_data_sheet(self, responses: List):
        """Create helper data sheet for statistical tests"""
        helper_sheet = self.workbook.create_sheet("Helper Data")
        
        # Define border style
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Title
        helper_sheet.merge_cells('A1:D1')
        title_cell = helper_sheet.cell(1, 1, "Helper Data for Statistical Tests")
        title_cell.font = Font(bold=True, size=14, color="FFFFFF")
        title_cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        title_cell.alignment = Alignment(horizontal="center", vertical="center")
        
        row = 3
        max_rows = len(responses) + 1
        
        # Question 3: Married and Unmarried scores
        helper_sheet.cell(row, 1, "QUESTION 3 - T-Test Helper Data").font = Font(bold=True, size=11, color="FFFFFF")
        helper_sheet.cell(row, 1).fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        helper_sheet.merge_cells(f'A{row}:D{row}')
        row += 1
        
        helper_sheet.cell(row, 1, "Married Scores").font = Font(bold=True)
        helper_sheet.cell(row, 1).fill = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")
        helper_sheet.cell(row, 1).border = thin_border
        helper_sheet.cell(row, 2, "Unmarried Scores").font = Font(bold=True)
        helper_sheet.cell(row, 2).fill = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")
        helper_sheet.cell(row, 2).border = thin_border
        q3_header_row = row
        row += 1
        
        q3_data_start = row
        for i in range(1, max_rows + 10):
            helper_sheet.cell(row, 1, f"=IF('Survey Data'!O{i+1}=\"Married\",'Survey Data'!C{i+1},\"\")").border = thin_border
            helper_sheet.cell(row, 2, f"=IF('Survey Data'!O{i+1}=\"Unmarried\",'Survey Data'!C{i+1},\"\")").border = thin_border
            row += 1
        q3_data_end = row - 1
        
        row += 3
        
        # Question 4: ANOVA helper data (Trendsetter by age group)
        helper_sheet.cell(row, 1, "QUESTION 4 - ANOVA Helper Data").font = Font(bold=True, size=11, color="FFFFFF")
        helper_sheet.cell(row, 1).fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        helper_sheet.merge_cells(f'A{row}:D{row}')
        row += 1
        
        helper_sheet.cell(row, 1, "18 to 34").font = Font(bold=True)
        helper_sheet.cell(row, 1).fill = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")
        helper_sheet.cell(row, 1).border = thin_border
        helper_sheet.cell(row, 2, "35 to 65").font = Font(bold=True)
        helper_sheet.cell(row, 2).fill = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")
        helper_sheet.cell(row, 2).border = thin_border
        helper_sheet.cell(row, 3, "65 and older").font = Font(bold=True)
        helper_sheet.cell(row, 3).fill = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")
        helper_sheet.cell(row, 3).border = thin_border
        q4_header_row = row
        row += 1
        
        q4_data_start = row
        age_groups = ["18 to 34", "35 to 65", "65 and older"]
        for i in range(1, max_rows + 10):
            for j, age_group in enumerate(age_groups):
                helper_sheet.cell(row, 1 + j, f"=IF('Survey Data'!P{i+1}=\"{age_group}\",'Survey Data'!J{i+1},\"\")").border = thin_border
            row += 1
        q4_data_end = row - 1
        
        row += 3
        
        # Question 5: Paired t-test helper data (Females Q5 and Q4)
        helper_sheet.cell(row, 1, "QUESTION 5 - Paired T-Test Helper Data").font = Font(bold=True, size=11, color="FFFFFF")
        helper_sheet.cell(row, 1).fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        helper_sheet.merge_cells(f'A{row}:D{row}')
        row += 1
        
        helper_sheet.cell(row, 1, "Q5 (Petrol Prices)").font = Font(bold=True)
        helper_sheet.cell(row, 1).fill = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")
        helper_sheet.cell(row, 1).border = thin_border
        helper_sheet.cell(row, 2, "Q4 (Alternatives)").font = Font(bold=True)
        helper_sheet.cell(row, 2).fill = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")
        helper_sheet.cell(row, 2).border = thin_border
        q5_header_row = row
        row += 1
        
        q5_data_start = row
        for i in range(1, max_rows + 10):
            helper_sheet.cell(row, 1, f"=IF('Survey Data'!N{i+1}=\"Female\",'Survey Data'!F{i+1},\"\")").border = thin_border
            helper_sheet.cell(row, 2, f"=IF('Survey Data'!N{i+1}=\"Female\",'Survey Data'!E{i+1},\"\")").border = thin_border
            row += 1
        q5_data_end = row - 1
        
        # Store ranges for use in statistical tests sheet
        self.helper_ranges = {
            'q3_married': f"'Helper Data'!A{q3_data_start}:A{q3_data_end}",
            'q3_unmarried': f"'Helper Data'!B{q3_data_start}:B{q3_data_end}",
            'q4_age1': f"'Helper Data'!A{q4_data_start}:A{q4_data_end}",
            'q4_age2': f"'Helper Data'!B{q4_data_start}:B{q4_data_end}",
            'q4_age3': f"'Helper Data'!C{q4_data_start}:C{q4_data_end}",
            'q5_petrol': f"'Helper Data'!A{q5_data_start}:A{q5_data_end}",
            'q5_alternatives': f"'Helper Data'!B{q5_data_start}:B{q5_data_end}"
        }
        
        # Adjust column widths
        helper_sheet.column_dimensions['A'].width = 20
        helper_sheet.column_dimensions['B'].width = 20
        helper_sheet.column_dimensions['C'].width = 20
        
        # Add survey footer
        self._add_survey_footer(helper_sheet)
    
    def _create_statistical_tests_sheet(self, responses: List):
        """Create statistical tests sheet for Questions 3, 4, and 5"""
        tests_sheet = self.workbook.create_sheet("Statistical Tests")
        
        # Define border style
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Title
        tests_sheet.merge_cells('A1:F1')
        title_cell = tests_sheet.cell(1, 1, "Statistical Tests - Questions 3, 4, and 5")
        title_cell.font = Font(bold=True, size=14, color="FFFFFF")
        title_cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        title_cell.alignment = Alignment(horizontal="center", vertical="center")
        
        row = 3
        
        # Question 3: T-Test for "Global warming is a real threat" between Married/Unmarried
        tests_sheet.cell(row, 1, "QUESTION 3").font = Font(bold=True, size=12, color="FFFFFF")
        tests_sheet.cell(row, 1).fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        tests_sheet.merge_cells(f'A{row}:F{row}')
        row += 1
        
        tests_sheet.cell(row, 1, "Research Question:").font = Font(bold=True)
        tests_sheet.cell(row, 2, "Is there a significant difference in the ratings of the statement 'Global warming is a real threat' between Married and Unmarried respondents?")
        tests_sheet.merge_cells(f'B{row}:F{row}')
        row += 1
        
        tests_sheet.cell(row, 1, "Null Hypothesis (H0):").font = Font(bold=True)
        tests_sheet.cell(row, 2, "μ_married = μ_unmarried (No significant difference in means)")
        tests_sheet.merge_cells(f'B{row}:F{row}')
        row += 1
        
        tests_sheet.cell(row, 1, "Alternative Hypothesis (H1):").font = Font(bold=True)
        tests_sheet.cell(row, 2, "μ_married ≠ μ_unmarried (Significant difference in means)")
        tests_sheet.merge_cells(f'B{row}:F{row}')
        row += 1
        
        tests_sheet.cell(row, 1, "Statistical Test:").font = Font(bold=True)
        tests_sheet.cell(row, 2, "Independent Samples T-Test")
        tests_sheet.merge_cells(f'B{row}:F{row}')
        row += 2
        
        # SUMMARY Section (like the examples)
        summary_start = row
        tests_sheet.cell(row, 1, "SUMMARY").font = Font(bold=True, size=11)
        tests_sheet.cell(row, 1).fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
        tests_sheet.merge_cells(f'A{row}:F{row}')
        row += 1
        
        # Summary table header with borders
        header_fill = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")
        tests_sheet.cell(row, 1, "Groups").font = Font(bold=True)
        tests_sheet.cell(row, 1).fill = header_fill
        tests_sheet.cell(row, 1).border = thin_border
        tests_sheet.cell(row, 2, "Count").font = Font(bold=True)
        tests_sheet.cell(row, 2).fill = header_fill
        tests_sheet.cell(row, 2).border = thin_border
        tests_sheet.cell(row, 3, "Sum").font = Font(bold=True)
        tests_sheet.cell(row, 3).fill = header_fill
        tests_sheet.cell(row, 3).border = thin_border
        tests_sheet.cell(row, 4, "Average").font = Font(bold=True)
        tests_sheet.cell(row, 4).fill = header_fill
        tests_sheet.cell(row, 4).border = thin_border
        tests_sheet.cell(row, 5, "Variance").font = Font(bold=True)
        tests_sheet.cell(row, 5).fill = header_fill
        tests_sheet.cell(row, 5).border = thin_border
        row += 1
        
        # Write statistics using helper data sheet references
        married_range = self.helper_ranges['q3_married']
        unmarried_range = self.helper_ranges['q3_unmarried']
        
        # Married row
        tests_sheet.cell(row, 1, "Married").border = thin_border
        tests_sheet.cell(row, 2, f"=COUNTIF({married_range},\">0\")").border = thin_border
        tests_sheet.cell(row, 2).number_format = '0.00'
        tests_sheet.cell(row, 3, f"=SUM({married_range})").border = thin_border
        tests_sheet.cell(row, 3).number_format = '0.00'
        tests_sheet.cell(row, 4, f"=AVERAGE({married_range})").border = thin_border
        tests_sheet.cell(row, 4).number_format = '0.00'
        tests_sheet.cell(row, 5, f"=VAR({married_range})").border = thin_border
        tests_sheet.cell(row, 5).number_format = '0.00'
        row += 1
        
        # Unmarried row
        tests_sheet.cell(row, 1, "Unmarried").border = thin_border
        tests_sheet.cell(row, 2, f"=COUNTIF({unmarried_range},\">0\")").border = thin_border
        tests_sheet.cell(row, 2).number_format = '0.00'
        tests_sheet.cell(row, 3, f"=SUM({unmarried_range})").border = thin_border
        tests_sheet.cell(row, 3).number_format = '0.00'
        tests_sheet.cell(row, 4, f"=AVERAGE({unmarried_range})").border = thin_border
        tests_sheet.cell(row, 4).number_format = '0.00'
        tests_sheet.cell(row, 5, f"=VAR({unmarried_range})").border = thin_border
        tests_sheet.cell(row, 5).number_format = '0.00'
        row += 2
        
        # T-Test Results Section
        p_value_row = row
        tests_sheet.cell(row, 1, "T-Test Results").font = Font(bold=True, size=11)
        tests_sheet.cell(row, 1).fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
        tests_sheet.merge_cells(f'A{row}:F{row}')
        row += 1
        
        tests_sheet.cell(row, 1, "p-value:").font = Font(bold=True)
        p_value_cell = f"B{row}"
        # Use TTEST (older function) for better compatibility, with fallback to T.TEST
        # Type 2 = Two-sample equal variance t-test
        tests_sheet.cell(row, 2, f"=IFERROR(TTEST({married_range},{unmarried_range},2,2),IFERROR(T.TEST({married_range},{unmarried_range},2,2),\"Error\"))")
        tests_sheet.cell(row, 2).number_format = '0.0000'
        row += 1
        tests_sheet.cell(row, 1, "Significance Level (α):").font = Font(bold=True)
        tests_sheet.cell(row, 2, "0.05")
        row += 1
        tests_sheet.cell(row, 1, "Conclusion:").font = Font(bold=True)
        tests_sheet.cell(row, 2, f'=IF({p_value_cell}<0.05,"Reject H0 - There is a statistically significant difference","Fail to reject H0 - No statistically significant difference")')
        tests_sheet.merge_cells(f'B{row}:F{row}')
        
        row += 3
        
        row += 3
        
        # Question 4: ANOVA for Trendsetter across age groups
        tests_sheet.cell(row, 1, "QUESTION 4").font = Font(bold=True, size=12, color="FFFFFF")
        tests_sheet.cell(row, 1).fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        tests_sheet.merge_cells(f'A{row}:F{row}')
        row += 1
        
        tests_sheet.cell(row, 1, "Research Question:").font = Font(bold=True)
        tests_sheet.cell(row, 2, "Do the mean scores of the personality description 'Trendsetter' differ between age groups?")
        tests_sheet.merge_cells(f'B{row}:F{row}')
        row += 1
        
        tests_sheet.cell(row, 1, "Null Hypothesis (H0):").font = Font(bold=True)
        tests_sheet.cell(row, 2, "μ_18-34 = μ_35-65 = μ_65+ (No significant difference in means across groups)")
        tests_sheet.merge_cells(f'B{row}:F{row}')
        row += 1
        
        tests_sheet.cell(row, 1, "Alternative Hypothesis (H1):").font = Font(bold=True)
        tests_sheet.cell(row, 2, "At least one group mean is significantly different")
        tests_sheet.merge_cells(f'B{row}:F{row}')
        row += 1
        
        tests_sheet.cell(row, 1, "Statistical Test:").font = Font(bold=True)
        tests_sheet.cell(row, 2, "One-Way ANOVA")
        tests_sheet.merge_cells(f'B{row}:F{row}')
        row += 2
        
        # SUMMARY Section (like ANOVA examples)
        tests_sheet.cell(row, 1, "SUMMARY").font = Font(bold=True, size=11)
        tests_sheet.cell(row, 1).fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
        tests_sheet.merge_cells(f'A{row}:F{row}')
        row += 1
        
        # Summary table header with borders
        header_fill = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")
        tests_sheet.cell(row, 1, "Groups").font = Font(bold=True)
        tests_sheet.cell(row, 1).fill = header_fill
        tests_sheet.cell(row, 1).border = thin_border
        tests_sheet.cell(row, 2, "Count").font = Font(bold=True)
        tests_sheet.cell(row, 2).fill = header_fill
        tests_sheet.cell(row, 2).border = thin_border
        tests_sheet.cell(row, 3, "Sum").font = Font(bold=True)
        tests_sheet.cell(row, 3).fill = header_fill
        tests_sheet.cell(row, 3).border = thin_border
        tests_sheet.cell(row, 4, "Average").font = Font(bold=True)
        tests_sheet.cell(row, 4).fill = header_fill
        tests_sheet.cell(row, 4).border = thin_border
        tests_sheet.cell(row, 5, "Variance").font = Font(bold=True)
        tests_sheet.cell(row, 5).fill = header_fill
        tests_sheet.cell(row, 5).border = thin_border
        row += 1
        
        # Write statistics using helper data sheet references
        age1_range = self.helper_ranges['q4_age1']
        age2_range = self.helper_ranges['q4_age2']
        age3_range = self.helper_ranges['q4_age3']
        
        # Age 18-34 row
        tests_sheet.cell(row, 1, "18 to 34").border = thin_border
        tests_sheet.cell(row, 2, f"=COUNTIF({age1_range},\">0\")").border = thin_border
        tests_sheet.cell(row, 2).number_format = '0.00'
        tests_sheet.cell(row, 3, f"=SUM({age1_range})").border = thin_border
        tests_sheet.cell(row, 3).number_format = '0.00'
        tests_sheet.cell(row, 4, f"=AVERAGE({age1_range})").border = thin_border
        tests_sheet.cell(row, 4).number_format = '0.00'
        tests_sheet.cell(row, 5, f"=VAR({age1_range})").border = thin_border
        tests_sheet.cell(row, 5).number_format = '0.00'
        row += 1
        
        # Age 35-65 row
        tests_sheet.cell(row, 1, "35 to 65").border = thin_border
        tests_sheet.cell(row, 2, f"=COUNTIF({age2_range},\">0\")").border = thin_border
        tests_sheet.cell(row, 2).number_format = '0.00'
        tests_sheet.cell(row, 3, f"=SUM({age2_range})").border = thin_border
        tests_sheet.cell(row, 3).number_format = '0.00'
        tests_sheet.cell(row, 4, f"=AVERAGE({age2_range})").border = thin_border
        tests_sheet.cell(row, 4).number_format = '0.00'
        tests_sheet.cell(row, 5, f"=VAR({age2_range})").border = thin_border
        tests_sheet.cell(row, 5).number_format = '0.00'
        row += 1
        
        # Age 65+ row
        tests_sheet.cell(row, 1, "65 and older").border = thin_border
        tests_sheet.cell(row, 2, f"=COUNTIF({age3_range},\">0\")").border = thin_border
        tests_sheet.cell(row, 2).number_format = '0.00'
        tests_sheet.cell(row, 3, f"=SUM({age3_range})").border = thin_border
        tests_sheet.cell(row, 3).number_format = '0.00'
        tests_sheet.cell(row, 4, f"=AVERAGE({age3_range})").border = thin_border
        tests_sheet.cell(row, 4).number_format = '0.00'
        tests_sheet.cell(row, 5, f"=VAR({age3_range})").border = thin_border
        tests_sheet.cell(row, 5).number_format = '0.00'
        row += 2
        
        # ANOVA Results Section - Automated Calculation using helper cells
        tests_sheet.cell(row, 1, "ANOVA Results").font = Font(bold=True, size=11)
        tests_sheet.cell(row, 1).fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
        tests_sheet.merge_cells(f'A{row}:F{row}')
        row += 1
        
        # Create helper calculation cells in columns G, H, I (will be hidden)
        calc_row = row
        # Row 1: n1, n2, n3
        tests_sheet.cell(calc_row, 7, f"=COUNTIF({age1_range},\">0\")")
        tests_sheet.cell(calc_row, 8, f"=COUNTIF({age2_range},\">0\")")
        tests_sheet.cell(calc_row, 9, f"=COUNTIF({age3_range},\">0\")")
        # Row 2: mean1, mean2, mean3
        tests_sheet.cell(calc_row + 1, 7, f"=AVERAGE({age1_range})")
        tests_sheet.cell(calc_row + 1, 8, f"=AVERAGE({age2_range})")
        tests_sheet.cell(calc_row + 1, 9, f"=AVERAGE({age3_range})")
        # Row 3: var1, var2, var3
        tests_sheet.cell(calc_row + 2, 7, f"=VAR({age1_range})")
        tests_sheet.cell(calc_row + 2, 8, f"=VAR({age2_range})")
        tests_sheet.cell(calc_row + 2, 9, f"=VAR({age3_range})")
        # Row 4: Grand mean numerator and denominator
        tests_sheet.cell(calc_row + 3, 7, f"=G{calc_row}*G{calc_row + 1}+H{calc_row}*H{calc_row + 1}+I{calc_row}*I{calc_row + 1}")
        tests_sheet.cell(calc_row + 3, 8, f"=G{calc_row}+H{calc_row}+I{calc_row}")
        tests_sheet.cell(calc_row + 3, 9, f"=G{calc_row + 3}/H{calc_row + 3}")  # Grand mean
        # Row 5: SSB, SSW, df_within
        tests_sheet.cell(calc_row + 4, 7, f"=G{calc_row}*(G{calc_row + 1}-I{calc_row + 3})^2+H{calc_row}*(H{calc_row + 1}-I{calc_row + 3})^2+I{calc_row}*(I{calc_row + 1}-I{calc_row + 3})^2")  # SSB
        tests_sheet.cell(calc_row + 4, 8, f"=(G{calc_row}-1)*G{calc_row + 2}+(H{calc_row}-1)*H{calc_row + 2}+(I{calc_row}-1)*I{calc_row + 2}")  # SSW
        tests_sheet.cell(calc_row + 4, 9, f"=G{calc_row}+H{calc_row}+I{calc_row}-3")  # df_within
        # Row 6: MSB, MSW, F-statistic
        tests_sheet.cell(calc_row + 5, 7, f"=G{calc_row + 4}/2")  # MSB
        tests_sheet.cell(calc_row + 5, 8, f"=H{calc_row + 4}/I{calc_row + 4}")  # MSW
        tests_sheet.cell(calc_row + 5, 9, f"=G{calc_row + 5}/H{calc_row + 5}")  # F-statistic
        
        # Display results (using helper cells)
        tests_sheet.cell(row, 1, "F-statistic:").font = Font(bold=True)
        tests_sheet.cell(row, 2, f"=I{calc_row + 5}")
        tests_sheet.cell(row, 2).number_format = '0.0000'
        row += 1
        tests_sheet.cell(row, 1, "df (Between):").font = Font(bold=True)
        tests_sheet.cell(row, 2, "2")
        row += 1
        tests_sheet.cell(row, 1, "df (Within):").font = Font(bold=True)
        tests_sheet.cell(row, 2, f"=I{calc_row + 4}")
        row += 1
        tests_sheet.cell(row, 1, "p-value:").font = Font(bold=True)
        p_value_cell = f"B{row}"
        # Use FDIST for compatibility (older Excel versions) - FDIST(x, df1, df2) = right-tail probability
        tests_sheet.cell(row, 2, f"=IFERROR(FDIST(I{calc_row + 5},2,I{calc_row + 4}),IFERROR(F.DIST.RT(I{calc_row + 5},2,I{calc_row + 4}),\"Error\"))")
        tests_sheet.cell(row, 2).number_format = '0.0000'
        row += 1
        tests_sheet.cell(row, 1, "Significance Level (α):").font = Font(bold=True)
        tests_sheet.cell(row, 2, "0.05")
        row += 1
        tests_sheet.cell(row, 1, "Conclusion:").font = Font(bold=True)
        tests_sheet.cell(row, 2, f'=IF({p_value_cell}<0.05,"Reject H0 - There is a statistically significant difference","Fail to reject H0 - No statistically significant difference")')
        tests_sheet.cell(row, 2).fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
        tests_sheet.merge_cells(f'B{row}:F{row}')
        
        # Hide helper calculation columns G, H, I
        tests_sheet.column_dimensions['G'].width = 0
        tests_sheet.column_dimensions['H'].width = 0
        tests_sheet.column_dimensions['I'].width = 0
        
        row += 3
        
        # Question 5: Paired T-Test for Females - Petrol Prices vs Alternatives
        tests_sheet.cell(row, 1, "QUESTION 5").font = Font(bold=True, size=12, color="FFFFFF")
        tests_sheet.cell(row, 1).fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        tests_sheet.merge_cells(f'A{row}:F{row}')
        row += 1
        
        tests_sheet.cell(row, 1, "Research Question:").font = Font(bold=True)
        tests_sheet.cell(row, 2, "Is there a statistically significant difference between females' beliefs of the level of petrol prices and females' views on the search for alternative fuel sources?")
        tests_sheet.merge_cells(f'B{row}:F{row}')
        row += 1
        
        tests_sheet.cell(row, 1, "Null Hypothesis (H0):").font = Font(bold=True)
        tests_sheet.cell(row, 2, "μ_petrol_prices = μ_alternatives (No significant difference)")
        tests_sheet.merge_cells(f'B{row}:F{row}')
        row += 1
        
        tests_sheet.cell(row, 1, "Alternative Hypothesis (H1):").font = Font(bold=True)
        tests_sheet.cell(row, 2, "μ_petrol_prices ≠ μ_alternatives (Significant difference)")
        tests_sheet.merge_cells(f'B{row}:F{row}')
        row += 1
        
        tests_sheet.cell(row, 1, "Statistical Test:").font = Font(bold=True)
        tests_sheet.cell(row, 2, "Paired Samples T-Test")
        tests_sheet.merge_cells(f'B{row}:F{row}')
        row += 2
        
        # SUMMARY Section
        tests_sheet.cell(row, 1, "SUMMARY").font = Font(bold=True, size=11)
        tests_sheet.cell(row, 1).fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
        tests_sheet.merge_cells(f'A{row}:F{row}')
        row += 1
        
        # Summary table header with borders
        header_fill = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")
        tests_sheet.cell(row, 1, "Groups").font = Font(bold=True)
        tests_sheet.cell(row, 1).fill = header_fill
        tests_sheet.cell(row, 1).border = thin_border
        tests_sheet.cell(row, 2, "Count").font = Font(bold=True)
        tests_sheet.cell(row, 2).fill = header_fill
        tests_sheet.cell(row, 2).border = thin_border
        tests_sheet.cell(row, 3, "Sum").font = Font(bold=True)
        tests_sheet.cell(row, 3).fill = header_fill
        tests_sheet.cell(row, 3).border = thin_border
        tests_sheet.cell(row, 4, "Average").font = Font(bold=True)
        tests_sheet.cell(row, 4).fill = header_fill
        tests_sheet.cell(row, 4).border = thin_border
        tests_sheet.cell(row, 5, "Variance").font = Font(bold=True)
        tests_sheet.cell(row, 5).fill = header_fill
        tests_sheet.cell(row, 5).border = thin_border
        row += 1
        
        # Write statistics using helper data sheet references
        q5_range = self.helper_ranges['q5_petrol']
        q4_range = self.helper_ranges['q5_alternatives']
        
        # Q5 (Petrol Prices) row
        tests_sheet.cell(row, 1, "Q5 (Petrol Prices)").border = thin_border
        tests_sheet.cell(row, 2, f"=COUNTIF({q5_range},\">0\")").border = thin_border
        tests_sheet.cell(row, 2).number_format = '0.00'
        tests_sheet.cell(row, 3, f"=SUM({q5_range})").border = thin_border
        tests_sheet.cell(row, 3).number_format = '0.00'
        tests_sheet.cell(row, 4, f"=AVERAGE({q5_range})").border = thin_border
        tests_sheet.cell(row, 4).number_format = '0.00'
        tests_sheet.cell(row, 5, f"=VAR({q5_range})").border = thin_border
        tests_sheet.cell(row, 5).number_format = '0.00'
        row += 1
        
        # Q4 (Alternatives) row
        tests_sheet.cell(row, 1, "Q4 (Alternatives)").border = thin_border
        tests_sheet.cell(row, 2, f"=COUNTIF({q4_range},\">0\")").border = thin_border
        tests_sheet.cell(row, 2).number_format = '0.00'
        tests_sheet.cell(row, 3, f"=SUM({q4_range})").border = thin_border
        tests_sheet.cell(row, 3).number_format = '0.00'
        tests_sheet.cell(row, 4, f"=AVERAGE({q4_range})").border = thin_border
        tests_sheet.cell(row, 4).number_format = '0.00'
        tests_sheet.cell(row, 5, f"=VAR({q4_range})").border = thin_border
        tests_sheet.cell(row, 5).number_format = '0.00'
        row += 2
        
        # Paired T-Test Results Section
        tests_sheet.cell(row, 1, "Paired T-Test Results").font = Font(bold=True, size=11)
        tests_sheet.cell(row, 1).fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
        tests_sheet.merge_cells(f'A{row}:F{row}')
        row += 1
        
        tests_sheet.cell(row, 1, "p-value:").font = Font(bold=True)
        p_value_cell = f"B{row}"
        # Use TTEST (older function) for better compatibility, with fallback to T.TEST
        # Type 1 = Paired two-sample t-test
        tests_sheet.cell(row, 2, f"=IFERROR(TTEST({q5_range},{q4_range},2,1),IFERROR(T.TEST({q5_range},{q4_range},2,1),\"Error\"))")
        tests_sheet.cell(row, 2).number_format = '0.0000'
        row += 1
        tests_sheet.cell(row, 1, "Significance Level (α):").font = Font(bold=True)
        tests_sheet.cell(row, 2, "0.05")
        row += 1
        tests_sheet.cell(row, 1, "Conclusion:").font = Font(bold=True)
        tests_sheet.cell(row, 2, f'=IF({p_value_cell}<0.05,"Reject H0 - There is a statistically significant difference","Fail to reject H0 - No statistically significant difference")')
        tests_sheet.merge_cells(f'B{row}:F{row}')
        
        # Adjust column widths for clean layout (no helper data in this sheet)
        tests_sheet.column_dimensions['A'].width = 20  # Statistic labels
        tests_sheet.column_dimensions['B'].width = 15  # Statistics values
        tests_sheet.column_dimensions['C'].width = 15  # Statistics values
        tests_sheet.column_dimensions['D'].width = 15  # Statistics values (ANOVA)
        tests_sheet.column_dimensions['E'].width = 50  # Merged cells
        tests_sheet.column_dimensions['F'].width = 50  # Merged cells
        
        # Add survey footer
        self._add_survey_footer(tests_sheet)
